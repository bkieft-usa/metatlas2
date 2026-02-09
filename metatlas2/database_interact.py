import pandas as pd
import numpy as np
import duckdb
import uuid
import sys
import os
import json
import time
import sys
from pathlib import Path
from typing import Dict
from tqdm.notebook import tqdm
from typing import Dict, List, Optional, Any, Tuple

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from contextlib import contextmanager

sys.path.append('/Users/BKieft/Metabolomics/metatlas2/metatlas2')
import lcmsruns_tools as lrt
import pubchem_retrieval as pcr
import load_tools as ldt
import logging_config as lcf
import rt_align_tools as rat

# Initialize logger properly at module level
logger = lcf.get_logger('database_interact')

@contextmanager
def get_db_connection(db_path: str):
    conn = duckdb.connect(str(db_path))
    try:
        yield conn
    finally:
        conn.close()

def get_files_by_type_from_db(project_db_path: str, file_type: str) -> pd.DataFrame:
    """
    Get files of a specific type from the project database.
    
    Args:
        project_db_path: Path to project database
        file_type: Type of files to retrieve ('qc', 'experimental', 'istd', 'injbl', 'exctrl')
    
    Returns:
        DataFrame with columns: file_path, chromatography, polarity, file_type
    """

    try:
        with get_db_connection(project_db_path) as conn:
            files_df = conn.execute("""
                SELECT file_path, chromatography, polarity, file_type 
                FROM lcmsruns 
                WHERE file_type = ?
                ORDER BY chromatography, polarity, file_path
            """, [file_type]).df()
        
        if files_df.empty:
            raise ValueError(f"No {file_type} files found in the project database")

        return files_df
    except Exception as e:
        logger.error(f"Error retrieving {file_type} files from database at {project_db_path}: {e}")
        return pd.DataFrame()

def get_atlas_compounds_table(database_path: str, atlas_uid: str, main_db_path: str = None) -> pd.DataFrame:
    """
    Extract all compound information for a given atlas UID from the database and return as a pandas DataFrame.
    Handles both main database (with mz_rt_references) and project database (with mz_rt_experimental) structures.
    Uses external database attachment for compound metadata when needed.
    """
    with get_db_connection(database_path) as conn:
    
        try:
            # Check if this is a project database by looking for mz_rt_experimental table
            is_project_db = conn.execute("""
                SELECT COUNT(*) 
                FROM information_schema.tables 
                WHERE table_name = 'mz_rt_experimental'
            """).fetchone()[0] > 0
        except Exception as e:
            logger.error(f"Error checking database type at {database_path}: {e}")
            return pd.DataFrame()

        # Attach main database if this is a project database and main_db_path provided
        try:
            if is_project_db and main_db_path:
                conn.execute(f"ATTACH '{main_db_path}' AS main_db")
                logger.info("Attached main database for compound metadata")
        except Exception as e:
            logger.error(f"Error attaching main database at {main_db_path}: {e}")

        try:
            if is_project_db:
                # Project database query - uses experimental data with external compound references
                query = """
                    SELECT
                        a.atlas_uid,
                        a.atlas_name,
                        a.atlas_description,
                        a.chromatography,
                        a.polarity,
                        aca.compound_uid,
                        COALESCE(main_db.compounds.name, '') AS compound_name,
                        COALESCE(main_db.compounds.inchi_key, '') AS inchi_key,
                        COALESCE(main_db.compounds.inchi, '') AS inchi,
                        mzrt_exp.adduct,
                        mzrt_exp.mz,
                        mzrt_exp.rt_peak,
                        mzrt_exp.rt_min,
                        mzrt_exp.rt_max,
                        mzrt_exp.mz_tolerance,
                        mzrt_exp.mz_rt_experimental_uid AS mz_rt_reference_uid,
                        mzrt_exp.rt_correction_applied,
                        mzrt_exp.rt_shift,
                        mzrt_exp.source_mz_rt_reference_uid
                    FROM atlases a
                    JOIN atlas_compound_associations aca ON a.atlas_uid = aca.atlas_uid
                    LEFT JOIN main_db.compounds ON aca.compound_uid = main_db.compounds.compound_uid
                    LEFT JOIN mz_rt_experimental mzrt_exp ON aca.mz_rt_reference_uid = mzrt_exp.mz_rt_experimental_uid
                    WHERE a.atlas_uid = ?
                    ORDER BY aca.association_order, mzrt_exp.rt_peak
                """
            else:
                # Main database query - uses reference data with full compound info
                query = """
                    SELECT
                        a.atlas_uid,
                        a.atlas_name,
                        a.atlas_description,
                        a.chromatography,
                        a.polarity,
                        c.compound_uid,
                        c.name AS compound_name,
                        c.inchi_key,
                        c.inchi,
                        mzrt.adduct,
                        mzrt.mz,
                        mzrt.rt_peak,
                        mzrt.rt_min,
                        mzrt.rt_max,
                        mzrt.mz_tolerance,
                        mzrt.mz_rt_reference_uid,
                        FALSE AS rt_correction_applied,
                        NULL AS rt_shift,
                        NULL AS source_mz_rt_reference_uid
                    FROM atlases a
                    JOIN atlas_compound_associations aca ON a.atlas_uid = aca.atlas_uid
                    JOIN compounds c ON aca.compound_uid = c.compound_uid
                    LEFT JOIN mz_rt_references mzrt ON aca.mz_rt_reference_uid = mzrt.mz_rt_reference_uid
                    WHERE a.atlas_uid = ?
                    ORDER BY aca.association_order, mzrt.rt_peak
                """

            df = conn.execute(query, [atlas_uid]).df()

        except Exception as e:
            logger.error(f"Error preparing query for atlas {atlas_uid}: {e}")
            return pd.DataFrame()
        
        # Detach main database if attached
        if is_project_db and main_db_path:
            conn.execute("DETACH main_db")

    if df.empty:
        logger.warning(f"No compounds found for atlas {atlas_uid}")
        return pd.DataFrame()
    else:
        db_type = "project" if is_project_db else "main"
        chromatography = ldt.detect_atlas_input_chromatography(df)
        polarity = ldt.detect_atlas_input_polarity(df)
        df['label'] = df['compound_name'] if 'compound_name' in df.columns else ''
        logger.info(f"Retrieved {len(df)} compounds from {db_type} database for atlas: {df['atlas_name'].iloc[0]} ({df['atlas_uid'].iloc[0]})")
        logger.debug(f"Atlas chromatography: {chromatography}, polarity: {polarity}")
        if is_project_db and df['rt_correction_applied'].any():
            logger.info("RT-corrected atlas detected with experimental data")

    return df

def check_rt_and_analysis_numbers(project_db_path: str, rt_alignment_number: int, analysis_number: int) -> None:
    """
    Check if the provided RT alignment number and analysis number already exist in the project database.
    If they do, raise an error to prevent overwriting existing data.
    """
    with get_db_connection(project_db_path) as conn:
        rt_exists = conn.execute("""
            SELECT COUNT(*) 
            FROM rt_alignment 
            WHERE rt_alignment_number = ?
        """, [rt_alignment_number]).fetchone()[0] > 0

        analysis_exists = conn.execute("""
            SELECT COUNT(*) 
            FROM mz_rt_experimental 
            WHERE analysis_number = ?
            AND rt_alignment_number = ?
        """, [analysis_number, rt_alignment_number]).fetchone()[0] > 0

    if rt_exists and analysis_exists:
        raise ValueError(f"RT alignment number {rt_alignment_number} and analysis number {analysis_number} already exists in the project database. Please increment one.")
    elif rt_exists and not analysis_exists:
        logger.info(f"Creating new analysis ({analysis_number}) for existing RT alignment ({rt_alignment_number}).")
    elif not rt_exists and analysis_exists:
        pass
    elif not rt_exists and not analysis_exists:
        logger.info(f"Creating new RT alignment ({rt_alignment_number}) and new analysis ({analysis_number}).")

    return

def create_project_database(project_db_path: str, rt_alignment_number: int, analysis_number: int) -> None:
    """
    Create project-specific database with required tables.
    Never overwrites existing databases - requires analyst to increment analysis number.
    """
    project_db_path = Path(project_db_path)
    project_db_path.parent.mkdir(parents=True, exist_ok=True)

    if project_db_path.exists():
        check_rt_and_analysis_numbers(project_db_path, rt_alignment_number, analysis_number)
        logger.info(f"Project database already exists at {project_db_path}. Proceeding with existing database and new RT alignment {rt_alignment_number} and/or analysis {analysis_number}.")

    with get_db_connection(project_db_path) as conn:
        _create_database_tables(conn, db_type="project")
    logger.info(f"Project database created at {project_db_path}")
    return

def create_metatlas_database(db_path: str, overwrite_existing: bool) -> None:
    """
    Create main metatlas database with required tables.
    """
    db_path = Path(db_path)
    db_path.parent.mkdir(parents=True, exist_ok=True)
    
    if overwrite_existing and db_path.exists():
        logger.warning("Overwriting existing main database (overwrite_existing=True)")
        db_path.unlink()
        logger.info(f"Deleted existing database at {db_path}")
    elif not overwrite_existing and db_path.exists():
        logger.warning(f"Database already exists at {db_path}. Use overwrite_existing=True to replace it.")
        return

    with get_db_connection(db_path) as conn:
        _create_database_tables(conn, db_type="main")

    logger.info(f"Main metatlas database created at {db_path}")
    return

def save_lcmsruns_to_db(
    project_db_path: str,
    project_name: str,
    project_lcmsruns_path: str,
    new_lcmsruns: str = False
) -> Dict:
    """
    Save LCMS run files to project database and return file paths grouped by chromatography/polarity/analysis type.
    Chromatography and polarity are inferred from filenames.
    If the lcmsruns table exists and has rows, do not overwrite unless new_lcmsruns is true
    """

    with get_db_connection(project_db_path) as conn:

        files_by_group = lrt.get_project_files(project_lcmsruns_path)
        prov = ldt.get_provenance()

        # Check if lcmsruns table exists
        table_exists = conn.execute(
            "SELECT COUNT(*) FROM information_schema.tables WHERE table_name = 'lcmsruns'"
        ).fetchone()[0] > 0

        # Check if lcmsruns table has rows
        has_rows = False
        if table_exists:
            has_rows = conn.execute("SELECT COUNT(*) FROM lcmsruns").fetchone()[0] > 0
            if has_rows:
                if not new_lcmsruns:
                    logger.info("LCMS runs already exist in the database and new_lcmsruns is False. Proceeding with these files.")
                    conn.close()
                    print_files_summary(files_by_group)
                    return files_by_group
                elif new_lcmsruns:
                    logger.info(f"LCMS runs already exist in the database for {project_name} and new_lcmsruns is True. Creating a new table...")
                    conn.execute("DELETE FROM lcmsruns")
            else:
                logger.info(f"No LCMS runs detected in project database for {project_name}. Creating a new table...")
        else:
            logger.info(f"No LCMS runs detected in project database for {project_name}. Creating a new table...")
        
        total_files = 0
        for chrom, pol_dict in files_by_group.items():
            for pol, analysis_dict in pol_dict.items():
                for file_type, file_list in analysis_dict.items():
                    for file_path in file_list:
                        filename = os.path.basename(file_path)
                        conn.execute(
                            "INSERT INTO lcmsruns VALUES (?, ?, ?, ?, ?, ?, ?)",
                            (
                                file_path,
                                filename,
                                file_type,
                                chrom,
                                pol,
                                prov["analyst"],
                                prov["timestamp"],
                            ),
                        )
                        total_files += 1

    logger.info(f"Saved {total_files} LCMS runs to database:")
    print_files_summary(files_by_group)

    return files_by_group

def print_files_summary(file_list: Dict[str, Any]):
    for chrom, pol_dict in file_list.items():
        logger.info(f"Chromatography: {chrom}")
        for pol, filetype_dict in pol_dict.items():
            logger.info(f"  Polarity: {pol}")
            for file_type, files_list in filetype_dict.items():
                logger.info(f"    {file_type}: {len(files_list)} files")

def list_available_atlases(db_path: str) -> pd.DataFrame:
    """List all available atlases with optional filtering."""

    try:
        with get_db_connection(db_path) as conn:
            conditions = ["1=1"]
            params = []
            
            query = f"""
            SELECT 
                a.atlas_uid,
                a.atlas_name,
                a.atlas_description,
                a.chromatography,
                a.polarity,
                a.last_modified,
                COUNT(aca.compound_uid) as compound_count
            FROM atlases a
            LEFT JOIN atlas_compound_associations aca ON a.atlas_uid = aca.atlas_uid
            WHERE {' AND '.join(conditions)}
            GROUP BY a.atlas_uid, a.atlas_name, a.atlas_description, 
                    a.chromatography, a.polarity, a.last_modified
            ORDER BY a.atlas_name
            """
            
            df = conn.execute(query, params).df()
        
        return df
    except:
        #logger.warning(f"Did not find any atlases in database at {db_path}")
        return pd.DataFrame()

def get_most_recent_QC_atlas_id(config: Dict, database_path: str = "main"):
    with get_db_connection(database_path) as conn:

        query = """
        SELECT *
        FROM atlases
        WHERE atlas_name LIKE '%QC%' OR atlas_description LIKE '%QC%'
        ORDER BY last_modified DESC
        LIMIT 1
        """
        df = conn.execute(query).df()

    if df.empty:
        logger.warning("No QC atlas found.")
        return None

    return df.iloc[0]['atlas_uid']

def validate_database(database_path: str, database_type: str = "main") -> None:
    """
    Validate the DuckDB database and print summary statistics.
    If database_path is "main", validates main DB structure.
    If database_path is a custom project DB, validates project DB structure.
    """
    logger.info("Database Validation:")

    db_path = Path(database_path)
    if not os.path.exists(db_path):
        logger.error(f"Database not found at: {db_path}")
        return

    with get_db_connection(db_path) as conn:
        if database_type == "main":
            # Main DB: compounds, references, atlases, associations
            compounds_count = conn.execute("SELECT COUNT(*) FROM compounds").fetchone()[0]
            references_count = conn.execute("SELECT COUNT(*) FROM mz_rt_references").fetchone()[0]
            atlases_count = conn.execute("SELECT COUNT(*) FROM atlases").fetchone()[0]
            atlas_compound_associations_count = conn.execute("SELECT COUNT(*) FROM atlas_compound_associations").fetchone()[0]
            atlas_info = list_available_atlases(db_path)

            method_combinations = conn.execute("""
                SELECT chromatography, polarity, COUNT(*) as reference_count 
                FROM mz_rt_references 
                GROUP BY chromatography, polarity
            """).fetchall()

            logger.info(f"   Compounds: {compounds_count}")
            logger.info(f"   RT/MZ References: {references_count}")
            logger.info(f"   Atlases: {atlases_count}")
            logger.info(f"   Atlas-Compound associations: {atlas_compound_associations_count}")

            if method_combinations:
                logger.info("   Method combinations:")
                for combo in method_combinations:
                    logger.info(f"      {combo[0]}/{combo[1]}: {combo[2]} references")

            if not atlas_info.empty:
                logger.info("   Available atlases:")
                for _, row in atlas_info.iterrows():
                    logger.info(f"      {row['atlas_uid']}")
                    logger.debug(f"            {row['atlas_name']}")
                    logger.debug(f"            {row['chromatography']} {row['polarity']}")
                    logger.debug(f"            {row['compound_count']} compounds")
                    logger.debug(f"            {row['last_modified']}")
        else:
            # Project DB: atlases, targeted_analysis, rt_alignment, mz_rt_experimental
            atlases_count = conn.execute("SELECT COUNT(*) FROM atlases").fetchone()[0]
            # Count unique analysis_uid values in targeted_analysis table
            #targeted_count = len(conn.execute("SELECT DISTINCT analysis_uid FROM targeted_analysis").fetchall())
            rt_alignment_count = conn.execute("SELECT COUNT(*) FROM rt_alignment").fetchone()[0]
            mz_rt_exp_count = conn.execute("SELECT COUNT(*) FROM mz_rt_experimental").fetchone()[0]

            logger.info(f"   Atlases: {atlases_count}")
            #logger.info(f"   Targeted analyses: {targeted_count}")
            logger.info(f"   RT alignment models: {rt_alignment_count}")
            logger.info(f"   Experimental RT/MZ entries: {mz_rt_exp_count}")

            # List atlases
            atlas_info = list_available_atlases(db_path)
            if not atlas_info.empty:
                logger.info("   Available atlases:")
                for _, row in atlas_info.iterrows():
                    logger.info(f"      {row['atlas_uid']}")
                    logger.debug(f"            {row['atlas_name']}")
                    logger.debug(f"            {row['chromatography']} {row['polarity']}")
                    logger.debug(f"            {row['compound_count']} compounds")
                    logger.debug(f"            {row['last_modified']}")

            # List targeted analyses
            # targeted_df = conn.execute("""
            #     SELECT analysis_uid, project_name, atlas_uid, COUNT(*) as compound_count
            #     FROM targeted_analysis
            #     GROUP BY analysis_uid, project_name, atlas_uid
            #     ORDER BY analysis_uid
            # """).df()
            # if not targeted_df.empty:
            #     logger.info("   Targeted analysis entries:")
            #     for _, row in targeted_df.iterrows():
            #         logger.info(f"      {row['analysis_uid']} ({row['project_name']}) - Atlas: {row['atlas_uid']} - {row['compound_count']} compounds")

            # List RT alignment models
            rt_df = conn.execute("""
                SELECT rt_alignment_uid, atlas_uid, model_type, polynomial_degree, r_squared, rmse, last_modified
                FROM rt_alignment
                ORDER BY last_modified DESC
            """).df()
            if not rt_df.empty:
                logger.info("   RT alignment models:")
                for _, row in rt_df.iterrows():
                    logger.info(f"      {row['rt_alignment_uid']} - Atlas: {row['atlas_uid']} - {row['model_type']} (deg={row['polynomial_degree']}, r2={row['r_squared']}, rmse={row['rmse']})")

            # List experimental RT/MZ entries summary
            exp_df = conn.execute("""
                SELECT chromatography, polarity, COUNT(*) as entry_count
                FROM mz_rt_experimental
                GROUP BY chromatography, polarity
            """).df()
            if not exp_df.empty:
                logger.info("   Experimental RT/MZ entries by method:")
                for _, row in exp_df.iterrows():
                    logger.info(f"      {row['chromatography']}/{row['polarity']}: {row['entry_count']} entries")

    return

def save_rt_alignment_model_to_db(qc_atlas_uid: str, project_db_path: Path, rt_alignment_number: int,
                                    best_model: dict, qc_files_info: pd.DataFrame, modeling_data: list) -> str:
    """Save RT alignment model to project database."""
    rt_alignment_uid = _generate_uid("rt_alignment")
    
    # Extract project name from path
    qc_files = qc_files_info['file_path'].tolist()
    project_db_path = Path(project_db_path)
    project_name = project_db_path.stem.replace('.duckdb', '')
    prov = ldt.get_provenance()

    model_metadata = {
        "qc_files": [os.path.basename(f) for f in qc_files],
        "compounds_used": [d.get('compound_uid', '') for d in modeling_data],
        "correction_timestamp": prov["timestamp"],
        "correction_method": "polynomial_qc_based",
        "analyst": prov["analyst"]
    }
    
    with get_db_connection(project_db_path) as conn:
        conn.execute("""
            INSERT INTO rt_alignment VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            rt_alignment_uid,
            project_name,
            rt_alignment_number,
            qc_atlas_uid,
            "polynomial",
            best_model['degree'],
            best_model['r2'],
            best_model['rmse'],
            json.dumps(best_model['coefficients'].tolist()),
            best_model['equation'],
            len(qc_files),
            len(modeling_data),
            prov["analyst"],
            prov["timestamp"],
            json.dumps(model_metadata)
        ))
    
    logger.info(f"RT alignment model saved to database with UID: {rt_alignment_uid}")
    return rt_alignment_uid

def add_compounds_to_db(input_df: pd.DataFrame, db_path: str, pubchem_cache_path: str, input_file_path: str = ""):
    """Add compounds and RT/MZ references to database using batch operations."""

    if not os.path.exists(db_path):
        logger.error(f"Database not found at {db_path}. Check path or create it first using create_metatlas_database().")
        raise FileNotFoundError(f"Database not found at {db_path}")

    unique_inchi_keys = input_df['inchi_key'].dropna().unique()
    logger.info(f"Adding {len(unique_inchi_keys)} compounds to database: {db_path}")
    
    pubchem_cache = pcr.load_or_create_pubchem_cache(pubchem_cache_path)
    prov = ldt.get_provenance()

    with get_db_connection(db_path) as conn:
        # Get existing compounds in one query
        existing_inchi_keys = set()
        existing_compounds_map = {}
        existing_result = conn.execute("SELECT inchi_key, compound_uid FROM compounds").fetchall()
        existing_inchi_keys = {row[0] for row in existing_result}
        existing_compounds_map = {row[0]: row[1] for row in existing_result}
        if existing_inchi_keys:
            logger.warning(f"Found {len(existing_inchi_keys)} existing compounds in database. Not creating duplicates.")

        # Prepare batch data
        compound_records = []
        reference_records = []
        compounds_created = 0
        compounds_skipped = 0

        logger.info("Preparing batch data...")
        for idx, row in tqdm(input_df.iterrows(), total=len(input_df), desc="Preparing compounds"):
            inchi_key = row.get('inchi_key')
            if pd.isna(inchi_key):
                continue
            
            chromatography = str(row.get('chromatography', 'Unknown'))
            polarity = str(row.get('polarity', 'Unknown'))
            compound_uid = None
            
            # Check if compound exists
            if inchi_key in existing_inchi_keys:
                compounds_skipped += 1
                compound_uid = existing_compounds_map.get(inchi_key)
            else:
                # Prepare new compound record
                compound_uid = _generate_uid("compound")
                compound_record = _prepare_compound_record(row, compound_uid, pubchem_cache, prov)
                compound_records.append(compound_record)
                compounds_created += 1
                
                # Add to existing sets for future duplicate checking in this session
                existing_inchi_keys.add(inchi_key)
                existing_compounds_map[inchi_key] = compound_uid
            
            # Prepare RT/MZ reference if data available and compound_uid exists
            if compound_uid and pd.notna(row.get('rt_peak')) and pd.notna(row.get('mz')):
                reference_record = _prepare_reference_record(
                    row, compound_uid, chromatography, polarity, input_file_path, prov
                )
                if reference_record:
                    reference_records.append(reference_record)

        # Batch insert compounds
        references_created = 0
        if compound_records:
            logger.info(f"Batch inserting {len(compound_records)} compounds...")
            conn.executemany("""
                INSERT INTO compounds VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, compound_records)

        # Check for existing references and batch insert if necessary
        if reference_records:
            logger.info(f"Checking for existing references and batch inserting if necessary...")
            # Get existing references to check for duplicates
            existing_refs = set()
            existing_ref_result = conn.execute("""
                SELECT compound_uid, chromatography, polarity, adduct 
                FROM mz_rt_references
            """).fetchall()
            existing_refs = {(row[0], row[1], row[2], row[3]) for row in existing_ref_result}

            # Filter out duplicates
            filtered_reference_records = []
            references_skipped = 0
            
            for record in reference_records:
                # Extract compound_uid, chromatography, polarity, adduct from record
                compound_uid = record[1]
                chromatography = record[8]
                polarity = record[9]
                adduct = record[7]
                
                ref_key = (compound_uid, chromatography, polarity, adduct)
                if ref_key not in existing_refs:
                    filtered_reference_records.append(record)
                    existing_refs.add(ref_key)  # Add to set to prevent duplicates within this batch
                    references_created += 1
                else:
                    references_skipped += 1

            # Batch insert filtered references
            if filtered_reference_records:
                conn.executemany("""
                    INSERT INTO mz_rt_references VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, filtered_reference_records)
        
        # Get final counts
        compounds_count = conn.execute("SELECT COUNT(*) FROM compounds").fetchone()[0]
        references_count = conn.execute("SELECT COUNT(*) FROM mz_rt_references").fetchone()[0]
    
    logger.info("Compounds added successfully!")
    logger.info(f"   Total compounds in database: {compounds_count}")
    logger.info(f"   New compounds created: {compounds_created}")
    if compounds_skipped > 0:
        logger.info(f"   Compounds skipped (already existed): {compounds_skipped}")
    logger.info(f"   Total RT/MZ references in database: {references_count}")
    logger.info(f"   New RT/MZ references created: {references_created}")
    if 'references_skipped' in locals() and references_skipped > 0:
        logger.info(f"   RT/MZ references skipped (duplicates): {references_skipped}")
    
    return

def _prepare_compound_record(row: pd.Series, compound_uid: str, pubchem_cache: Dict, prov: Dict) -> tuple:
    """Prepare a compound record tuple for batch insertion."""
    # Extract compound data from row
    name = str(row.get('label', row.get('compound_name', 'Unknown')))
    inchi_key = str(row.get('inchi_key', ''))
    pubchem_cache_data = pubchem_cache.get(inchi_key, {})

    inchi = str(pubchem_cache_data.get('inchi', '')) if pd.notna(pubchem_cache_data.get('inchi')) else None
    smiles = str(pubchem_cache_data.get('smiles', '')) if pd.notna(pubchem_cache_data.get('smiles')) else None
    formula = str(pubchem_cache_data.get('formula', '')) if pd.notna(pubchem_cache_data.get('formula')) else None
    mono_isotopic_molecular_weight = pubchem_cache_data.get('mono_isotopic_molecular_weight') if pd.notna(pubchem_cache_data.get('mono_isotopic_molecular_weight')) else None
    iupac_name = str(pubchem_cache_data.get('iupac_name', '')) if pd.notna(pubchem_cache_data.get('iupac_name')) else None
    pubchem_cid = str(pubchem_cache_data.get('pubchem_cid', '')) if pd.notna(pubchem_cache_data.get('pubchem_cid')) else None
    cas_number = str(pubchem_cache_data.get('cas_number', '')) if pd.notna(pubchem_cache_data.get('cas_number')) else None
    synonyms_val = pubchem_cache_data.get('synonyms', '')
    if synonyms_val is None or synonyms_val == '':
        synonyms = None
    elif isinstance(synonyms_val, (list, tuple, np.ndarray)):
        synonyms = '; '.join(map(str, synonyms_val))
    else:
        synonyms = str(synonyms_val)
    compound_classes = str(row.get('compound_classes', '')) if pd.notna(row.get('compound_classes')) else None
    compound_pathways = str(row.get('compound_pathways', '')) if pd.notna(row.get('compound_pathways')) else None
    compound_tags = str(row.get('compound_tags', '')) if pd.notna(row.get('compound_tags')) else None

    return (
        compound_uid,
        name,
        inchi_key,
        inchi,
        smiles,
        formula,
        compound_classes,
        compound_pathways,
        compound_tags,
        mono_isotopic_molecular_weight,
        iupac_name,
        pubchem_cid,
        cas_number,
        synonyms,
        prov["analyst"],
        prov["timestamp"]
    )

def _prepare_reference_record(row: pd.Series, compound_uid: str, chromatography: str, 
                            polarity: str, input_file_path: str, prov: Dict) -> tuple:
    """Prepare a reference record tuple for batch insertion."""
    try:
        uid = _generate_uid("mz_rt_reference")
        
        # Extract RT/MZ data
        rt_peak = float(row.get('rt_peak'))
        rt_min = float(row.get('rt_min', rt_peak - 0.5))
        rt_max = float(row.get('rt_max', rt_peak + 0.5))
        mz = float(row.get('mz'))
        mz_tolerance = float(row.get('mz_tolerance', 5.0))
        adduct = str(row.get('adduct', ''))
        
        confidence = str(row.get('confidence', 'Unknown')) if pd.notna(row.get('confidence')) else 'Unknown'
        source = input_file_path if input_file_path else 'Unknown'
        
        return (
            uid,
            compound_uid,
            rt_peak,
            rt_min,
            rt_max,
            mz,
            mz_tolerance,
            adduct,
            chromatography,
            polarity,
            confidence,
            source,
            prov["analyst"],
            prov["timestamp"]
        )
    except (ValueError, TypeError) as e:
        # Skip records with invalid data
        return None

def _prepare_compound_record_from_dict(compound_data: Dict) -> Optional[Tuple]:
    """Prepare compound record from dictionary data."""
    try:
        prov = ldt.get_provenance()
        
        # Use provided compound_uid (should be set by calling function)
        compound_uid = compound_data.get('compound_uid')
        if not compound_uid:
            logger.error("Compound data missing compound_uid - this should be set by calling function")
            return None
        
        # Ensure we have a valid inchi_key
        inchi_key = compound_data.get('inchi_key', '')
        if not inchi_key:
            logger.warning("Compound data missing inchi_key - skipping")
            return None
        
        return (
            compound_uid,
            compound_data.get('name', 'Unknown'),
            inchi_key,
            compound_data.get('inchi'),
            compound_data.get('smiles'),
            compound_data.get('formula'),
            compound_data.get('compound_classes'),
            compound_data.get('compound_pathways'),
            compound_data.get('compound_tags'),
            compound_data.get('mono_isotopic_molecular_weight'),
            compound_data.get('iupac_name'),
            compound_data.get('pubchem_cid'),
            compound_data.get('cas_number'),
            compound_data.get('synonyms'),
            prov["analyst"],
            prov["timestamp"]
        )
    except Exception as e:
        logger.error(f"Error preparing compound record: {e}")
        return None

def _prepare_reference_record_from_dict(reference_data: Dict) -> Optional[Tuple]:
    """Prepare reference record from dictionary data."""
    try:
        prov = ldt.get_provenance()
        
        return (
            _generate_uid("mz_rt_reference"),
            reference_data.get('compound_uid'),
            float(reference_data.get('rt_peak')),
            float(reference_data.get('rt_min')),
            float(reference_data.get('rt_max')),
            float(reference_data.get('mz')),
            float(reference_data.get('mz_tolerance', 5.0)),
            reference_data.get('adduct', ''),
            reference_data.get('chromatography'),
            reference_data.get('polarity'),
            reference_data.get('confidence', 'Unknown'),
            reference_data.get('source', 'Unknown'),
            prov["analyst"],
            prov["timestamp"]
        )
    except (ValueError, TypeError) as e:
        logger.error(f"Error preparing reference record: {e}")
        return None

# def generate_targeted_analysis_summary(project_db_path: str, config: Dict, analysis_uid: str) -> pd.DataFrame:
#     """
#     Generate summary of targeted analysis results from database.
    
#     Args:
#         project_db_path: Path to project database
#         config: Configuration dictionary
#         analysis_uid: UID of targeted analysis
    
#     Returns:
#         DataFrame with analysis summary
#     """
#     with get_db_connection(project_db_path) as conn:
#         # Get targeted analysis data
#         query = """
#             SELECT 
#                 ta.*,
#                 c.formula,
#                 c.mono_isotopic_molecular_weight
#             FROM targeted_analysis ta
#             LEFT JOIN compounds c ON ta.compound_uid = c.compound_uid
#             WHERE ta.analysis_uid = ?
#             ORDER BY ta.pre_rt_peak
#         """
        
#         try:
#             df = conn.execute(query, [analysis_uid]).df()
#         except Exception as e:
#             # Fallback query if compounds table doesn't exist in project DB
#             logger.warning(f"Could not join with compounds table: {e}")
#             query = """
#                 SELECT * FROM targeted_analysis 
#                 WHERE analysis_uid = ?
#                 ORDER BY pre_rt_peak
#             """
#             df = conn.execute(query, [analysis_uid]).df()
    
#     if df.empty:
#         logger.warning(f"No targeted analysis results found for analysis_uid: {analysis_uid}")
#     else:
#         logger.info(f"Retrieved {len(df)} compounds from targeted analysis {analysis_uid}")
    
#     return df

# def _calculate_msms_quality_score_from_notes(row: pd.Series) -> float:
#     """Calculate MS/MS quality score from analyst notes."""
#     ms2_notes = str(row.get('ms2_notes', 'no selection')).lower()
    
#     # Extract numeric score from notes if present
#     if '1.0' in ms2_notes:
#         return 3.0
#     elif '0.5' in ms2_notes:
#         return 1.5
#     elif '0.0' in ms2_notes:
#         return 0.0
#     elif '-1.0' in ms2_notes:
#         return 0.0
#     elif 'no selection' in ms2_notes:
#         return 0.0
#     else:
#         # Fallback to score if available
#         score = row.get('best_ms2_score', 0.0)
#         if pd.notna(score) and score > 0:
#             return min(3.0, float(score) * 3.0)
#         return 0.0

# def _calculate_mz_quality_score(row: pd.Series) -> float:
#     """Calculate m/z quality score based on PPM error."""
#     ppm_error = row.get('best_eic_ppm_error', None)
    
#     if pd.isna(ppm_error) or ppm_error is None:
#         return 0.0
    
#     ppm_error = abs(float(ppm_error))
    
#     # Score based on PPM error
#     if ppm_error <= 2.0:
#         return 2.0
#     elif ppm_error <= 5.0:
#         return 1.5
#     elif ppm_error <= 10.0:
#         return 1.0
#     elif ppm_error <= 20.0:
#         return 0.5
#     else:
#         return 0.0

# def _calculate_rt_quality_score(row: pd.Series) -> float:
#     """Calculate RT quality score based on RT error."""
#     rt_error = row.get('best_eic_rt_error', None)
    
#     if pd.isna(rt_error) or rt_error is None:
#         return 0.0
    
#     rt_error = abs(float(rt_error))
    
#     # Score based on RT error (in minutes)
#     if rt_error <= 0.1:
#         return 2.0
#     elif rt_error <= 0.2:
#         return 1.5
#     elif rt_error <= 0.5:
#         return 1.0
#     elif rt_error <= 1.0:
#         return 0.5
#     else:
#         return 0.0

# def _determine_msi_level(msms_quality: float, mz_quality: float, rt_quality: float) -> str:
#     """Determine MSI identification level based on quality scores."""
#     total_score = msms_quality + mz_quality + rt_quality
    
#     if total_score >= 6.0 and msms_quality >= 2.0:
#         return "MSI Level 2"
#     elif total_score >= 4.0 and mz_quality >= 1.0:
#         return "MSI Level 3"
#     elif total_score >= 2.0:
#         return "MSI Level 4"
#     else:
#         return "MSI Level 5"

# def _create_empty_summary_from_atlas(atlas_dataframe: pd.DataFrame, analysis_uid: str, config: Dict) -> pd.DataFrame:
#     """Create empty summary rows for all atlas compounds."""
#     empty_rows = []
    
#     for _, row in atlas_dataframe.iterrows():
#         empty_row = {
#             'analysis_uid': analysis_uid,
#             'compound_uid': row.get('compound_uid', ''),
#             'inchi_key': row.get('inchi_key', ''),
#             'compound_name': row.get('compound_name', row.get('label', '')),
#             'formula': row.get('formula', ''),
#             'mono_isotopic_molecular_weight': row.get('mono_isotopic_molecular_weight', ''),
#             'adduct': row.get('adduct', ''),
#             'pre_rt_peak': row.get('rt_peak', 0.0),
#             'pre_rt_min': row.get('rt_min', 0.0),
#             'pre_rt_max': row.get('rt_max', 0.0),
#             'pre_mz': row.get('mz', 0.0),
#             'ms1_notes': 'keep',
#             'ms2_notes': 'no selection',
#             'analyst_notes': '',
#             'identification_notes': '',
#             # Empty detection fields
#             'best_eic_intensity': 0.0,
#             'best_eic_file': '',
#             'best_eic_rt': 0.0,
#             'best_eic_mz': 0.0,
#             'best_eic_ppm_error': 0.0,
#             'best_eic_rt_error': 0.0,
#             'best_ms2_file': '',
#             'best_ms2_score': 0.0,
#             'best_ms2_num_matches': 0,
#             'best_ms2_matched_fragments': '',
#             'post_rt_peak': row.get('rt_peak', 0.0),
#             'post_rt_min': row.get('rt_min', 0.0),
#             'post_rt_max': row.get('rt_max', 0.0)
#         }
#         empty_rows.append(empty_row)
    
#     return pd.DataFrame(empty_rows)

# def _add_missing_compounds_to_summary(base_summary: pd.DataFrame, atlas_dataframe: pd.DataFrame, 
#                                      analysis_uid: str, config: Dict) -> pd.DataFrame:
#     """Add missing compounds from atlas to summary as empty rows."""
#     # Find compounds in atlas but not in summary
#     atlas_inchi_keys = set(atlas_dataframe['inchi_key'].unique())
#     summary_inchi_keys = set(base_summary['inchi_key'].unique())
#     missing_inchi_keys = atlas_inchi_keys - summary_inchi_keys
    
#     if missing_inchi_keys:
#         logger.info(f"Adding {len(missing_inchi_keys)} missing compounds to report")
        
#         missing_compounds = atlas_dataframe[atlas_dataframe['inchi_key'].isin(missing_inchi_keys)]
#         empty_summary = _create_empty_summary_from_atlas(missing_compounds, analysis_uid, config)
        
#         # Combine with existing summary
#         combined_summary = pd.concat([base_summary, empty_summary], ignore_index=True)
#         return combined_summary
    
#     return base_summary

# def _save_report_with_grouped_headers(report_df: pd.DataFrame, output_path: str, atlas_uid: str):
#     """Save report to Excel with grouped headers."""
#     try:
#         import openpyxl
#         from openpyxl.utils.dataframe import dataframe_to_rows
#         from openpyxl.styles import Font, Alignment, PatternFill
#         from openpyxl.utils import get_column_letter
        
#         # Create workbook and worksheet
#         wb = openpyxl.Workbook()
#         ws = wb.active
#         ws.title = "Targeted Analysis Report"
        
#         # Define column groups and their headers
#         column_groups = [
#             ('Compound Information', ['index', 'identified_metabolite', 'label', 'isomer_compound', 'isomer_inchi_keys', 'formula', 'polarity', 'mono_isotopic_molecular_weight', 'inchi_key', 'adduct']),
#             ('Quality Scores', ['msms_quality', 'mz_quality', 'rt_quality', 'total_score', 'msi_level']),
#             ('Analyst Annotations', ['ms1_notes', 'ms2_notes', 'analyst_notes', 'identification_notes']),
#             ('Detection Results', ['max_intensity', 'max_intensity_file', 'max_intensity_rt', 'best_msms_file', 'best_msms_rt', 'best_msms_num_matching_ions', 'best_msms_matching_ions', 'best_msms_score']),
#             ('Measurement Accuracy', ['mz_theoretical', 'mz_measured', 'mz_error', 'rt_peak_theoretical', 'rt_peak_measured', 'rt_min_measured', 'rt_max_measured', 'rt_error'])
#         ]
        
#         # Add title and metadata
#         ws['A1'] = f"Targeted Analysis Report - Atlas: {atlas_uid}"
#         ws['A1'].font = Font(bold=True, size=14)
#         ws['A2'] = f"Generated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}"
#         ws['A3'] = f"Total Compounds: {len(report_df)}"
        
#         # Start data at row 5
#         start_row = 5
#         current_col = 1
        
#         # Create grouped headers
#         for group_name, columns in column_groups:
#             # Check which columns exist in the dataframe
#             existing_columns = [col for col in columns if col in report_df.columns]
#             if not existing_columns:
#                 continue
            
#             # Group header
#             start_col = current_col
#             end_col = current_col + len(existing_columns) - 1
            
#             if start_col == end_col:
#                 ws.cell(row=start_row, column=start_col).value = group_name
#             else:
#                 ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
#                 ws.cell(row=start_row, column=start_col).value = group_name
            
#             ws.cell(row=start_row, column=start_col).font = Font(bold=True)
#             ws.cell(row=start_row, column=start_col).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
#             ws.cell(row=start_row, column=start_col).alignment = Alignment(horizontal="center")
            
#             # Column headers
#             for i, col in enumerate(existing_columns):
#                 ws.cell(row=start_row + 1, column=current_col + i).value = col
#                 ws.cell(row=start_row + 1, column=current_col + i).font = Font(bold=True)
            
#             current_col += len(existing_columns)
        
#         # Add data rows
#         data_start_row = start_row + 2
#         for row_idx, (_, row) in enumerate(report_df.iterrows()):
#             current_col = 1
#             for group_name, columns in column_groups:
#                 existing_columns = [col for col in columns if col in report_df.columns]
#                 for col in existing_columns:
#                     ws.cell(row=data_start_row + row_idx, column=current_col).value = row[col]
#                     current_col += 1
        
#         # Auto-adjust column widths
#         for col in ws.columns:
#             max_length = 0
#             column = col[0].column_letter
#             for cell in col:
#                 try:
#                     if len(str(cell.value)) > max_length:
#                         max_length = len(str(cell.value))
#                 except:
#                     pass
#             adjusted_width = min(max_length + 2, 50)
#             ws.column_dimensions[column].width = adjusted_width
        
#         # Save workbook
#         wb.save(output_path)
        
#     except ImportError:
#         logger.warning("openpyxl not available, saving as CSV instead")
#         report_df.to_csv(output_path.replace('.xlsx', '.csv'), index=False)
#     except Exception as e:
#         logger.error(f"Error saving Excel file: {e}")
#         # Fallback to CSV
#         report_df.to_csv(output_path.replace('.xlsx', '.csv'), index=False)

# def validate_targeted_analysis_data(project_db_path: str, analysis_uid: str = None) -> bool:
#     """
#     Validate targeted analysis results in database.
#     """
#     with get_db_connection(project_db_path) as conn:
        
#         # Get latest analysis if no specific UID provided
#         if analysis_uid:
#             condition = "WHERE analysis_uid = ?"
#             params = [analysis_uid]
#         else:
#             # Get most recent analysis
#             latest_result = conn.execute("""
#                 SELECT analysis_uid FROM targeted_analysis 
#                 ORDER BY analysis_timestamp DESC LIMIT 1
#             """).fetchone()
#             if not latest_result:
#                 return False
#             analysis_uid = latest_result[0]
#             condition = "WHERE analysis_uid = ?"
#             params = [analysis_uid]
        
#         # Validation queries
#         validation_query = f"""
#         SELECT 
#             COUNT(*) as total_records,
#             COUNT(CASE WHEN pre_rt_peak IS NOT NULL THEN 1 END) as records_with_rt,
#             COUNT(CASE WHEN pre_mz IS NOT NULL THEN 1 END) as records_with_mz,
#             COUNT(CASE WHEN best_eic_file IS NOT NULL AND best_eic_file != '' THEN 1 END) as records_with_eic,
#             COUNT(CASE WHEN avg_eic_rt IS NOT NULL THEN 1 END) as records_with_avg_eic_rt,
#             COUNT(CASE WHEN ms2_files_with_data > 0 THEN 1 END) as records_with_ms2,
#             COUNT(CASE WHEN is_rt_modified = true THEN 1 END) as modified_records
#         FROM targeted_analysis
#         {condition}
#         """
        
#         result = conn.execute(validation_query, params).fetchone()
    
#     if result and result[0] > 0:  # At least one record exists
#         total_records = result[0]
#         logger.info(f"Validation successful for analysis {analysis_uid}:")
#         logger.info(f"  Total records: {total_records}")
#         logger.info(f"  Records with RT data: {result[1]}")
#         logger.info(f"  Records with m/z data: {result[2]}")
#         logger.info(f"  Records with EIC data: {result[3]}")
#         logger.info(f"  Records with MS2 data: {result[5]}")
#         logger.info(f"  Modified records: {result[6]}")
#         return True
#     else:
#         logger.error(f"Validation failed for analysis {analysis_uid}: No records found")
#         return False

# def clone_and_modify_atlas(source_db_path: str, dest_db_path: str, source_atlas_uid: str,
#                           config: Dict, compound_updates: Dict[str, Dict],
#                           use_experimental_table: bool = True,
#                           new_atlas_description: str = "Modified Atlas") -> str:
#     """
#     Clone an atlas and apply compound modifications.
    
#     Args:
#         source_db_path: Path to source database
#         dest_db_path: Path to destination database  
#         source_atlas_uid: UID of source atlas to clone
#         config: Configuration dictionary
#         compound_updates: Dict mapping compound_uid to update dictionary
#         use_experimental_table: Whether to use experimental or reference table
#         new_atlas_description: Description for new atlas
    
#     Returns:
#         UID of new atlas
#     """
#     logger.info(f"Cloning atlas {source_atlas_uid} with {len(compound_updates)} modifications...")
    
#     # Generate new atlas UID
#     new_atlas_uid = _generate_uid("analyzed_atlas")
#     prov = ldt.get_provenance()
    
#     with get_db_connection(source_db_path) as source_conn:
#         with get_db_connection(dest_db_path) as dest_conn:
            
#             # Get source atlas metadata
#             source_atlas = source_conn.execute("""
#                 SELECT atlas_name, chromatography, polarity 
#                 FROM atlases WHERE atlas_uid = ?
#             """, [source_atlas_uid]).fetchone()
            
#             if not source_atlas:
#                 raise ValueError(f"Source atlas {source_atlas_uid} not found")
            
#             atlas_name, chromatography, polarity = source_atlas
#             new_atlas_name = f"{atlas_name} - {new_atlas_description}"
            
#             # Create new atlas
#             dest_conn.execute("""
#                 INSERT INTO atlases VALUES (?, ?, ?, ?, ?, ?, ?)
#             """, (
#                 new_atlas_uid,
#                 new_atlas_name,
#                 new_atlas_description,
#                 chromatography,
#                 polarity,
#                 prov["analyst"],
#                 prov["timestamp"]
#             ))
            
#             # Get source compounds and associations
#             source_data = source_conn.execute("""
#                 SELECT 
#                     aca.compound_uid,
#                     aca.association_order,
#                     c.name as compound_name,
#                     c.inchi_key,
#                     mzrt.rt_peak,
#                     mzrt.rt_min,
#                     mzrt.rt_max,
#                     mzrt.mz,
#                     mzrt.mz_tolerance,
#                     mzrt.adduct,
#                     mzrt.chromatography,
#                     mzrt.polarity
#                 FROM atlas_compound_associations aca
#                 LEFT JOIN compounds c ON aca.compound_uid = c.compound_uid
#                 LEFT JOIN mz_rt_experimental mzrt ON aca.mz_rt_reference_uid = mzrt.mz_rt_experimental_uid
#                 WHERE aca.atlas_uid = ?
#                 ORDER BY aca.association_order
#             """, [source_atlas_uid]).fetchall()
            
#             # Create modified experimental entries and associations
#             for row in source_data:
#                 compound_uid = row[0]
#                 association_order = row[1]
                
#                 # Apply modifications if they exist
#                 if compound_uid in compound_updates:
#                     updates = compound_updates[compound_uid]
#                     rt_peak = updates.get('rt_peak', row[4])
#                     rt_min = updates.get('rt_min', row[5]) 
#                     rt_max = updates.get('rt_max', row[6])
#                     ms1_notes = updates.get('ms1_notes', 'keep')
#                     ms2_notes = updates.get('ms2_notes', 'no selection')
#                 else:
#                     rt_peak = row[4]
#                     rt_min = row[5]
#                     rt_max = row[6]
#                     ms1_notes = 'keep'
#                     ms2_notes = 'no selection'
                
#                 # Create new experimental entry
#                 exp_uid = _generate_uid("mz_rt_experimental")
#                 dest_conn.execute("""
#                     INSERT INTO mz_rt_experimental VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
#                 """, (
#                     exp_uid,
#                     compound_uid,
#                     rt_peak,
#                     rt_min,
#                     rt_max,
#                     ms1_notes,
#                     ms2_notes,
#                     row[7],  # mz
#                     row[8],  # mz_tolerance
#                     row[9],  # adduct
#                     row[10], # chromatography
#                     row[11], # polarity
#                     True,    # rt_correction_applied
#                     0.0,     # rt_shift
#                     None,    # source_mz_rt_reference_uid
#                     prov["analyst"],
#                     prov["timestamp"]
#                 ))
                
#                 # Create association
#                 assoc_uid = _generate_uid("association")
#                 dest_conn.execute("""
#                     INSERT INTO atlas_compound_associations VALUES (?, ?, ?, ?, ?, ?, ?)
#                 """, (
#                     assoc_uid,
#                     new_atlas_uid,
#                     compound_uid,
#                     exp_uid,
#                     association_order,
#                     prov["analyst"],
#                     prov["timestamp"]
#                 ))
    
#     logger.info(f"Created new atlas: {new_atlas_uid}")
#     return new_atlas_uid

def get_atlas_metadata_from_db(db_path: str, atlas_uid: str, validation: bool = False) -> pd.DataFrame:
    """
    Retrieve atlas metadata from database for a specific atlas.
    Returns DataFrame with atlas information including name, description, chromatography, polarity, etc.
    """
    
    try:
        with get_db_connection(db_path) as conn: 
            query = """
            SELECT 
                atlas_uid,
                atlas_name,
                atlas_description,
                chromatography,
                polarity,
                created_by,
                last_modified
            FROM atlases 
            WHERE atlas_uid = ?
            """
            
            df = conn.execute(query, [atlas_uid]).df()
        
        if df.empty:
            #logger.warning(f"No atlas found with UID: {atlas_uid}")
            return pd.DataFrame()
        else:
            if validation is False:
                logger.info(f"Retrieved atlas metadata for: {df['atlas_name'].iloc[0]}")
            return df
    except:
        #logger.warning(f"Did not find atlas {atlas_uid} in database {db_path}")
        return pd.DataFrame()

def get_rt_correction_table_entry(db_path: Path, atlas_uid: str) -> Optional[dict]:
    """
    Retrieve the RT correction table entry for a given atlas UID from the rt_alignment table.
    Returns a dictionary with the entry if found, else None.
    """

    with get_db_connection(db_path) as conn:

        query = """
            SELECT *
            FROM rt_alignment
            WHERE atlas_uid = ?
            ORDER BY last_modified DESC
            LIMIT 1
        """
        result = conn.execute(query, [atlas_uid]).fetchone()
        columns = [desc[0] for desc in conn.description] if result else []

    if result and columns:
        rt_dict = dict(zip(columns, result))
        rt_df = pd.DataFrame([rt_dict])
        return rt_df
    else:
        logger.warning(f"No RT correction entry found for atlas UID: {atlas_uid}")
        return None

def _generate_uid(entity_type: str, decorator: str = None) -> str:
    """Generate a unique identifier for database entities."""
    if entity_type == "atlas":
        return f"atl-raw-{decorator}-{uuid.uuid4().hex[:32]}" if decorator else f"atl-raw-{uuid.uuid4().hex[:32]}"
    elif entity_type == "rt_atlas":
        return f"atl-rta-{decorator}-{uuid.uuid4().hex[:32]}" if decorator else f"atl-rta-{uuid.uuid4().hex[:32]}"
    elif entity_type == "analyzed_atlas":
        return f"atl-tga-{decorator}-{uuid.uuid4().hex[:32]}" if decorator else f"atl-tga-{uuid.uuid4().hex[:32]}"
    elif entity_type == "mz_rt_experimental":
        return f"mzrt-exp-{uuid.uuid4().hex[:32]}"
    elif entity_type == "compound":
        return f"cmp-{uuid.uuid4().hex[:32]}"
    elif entity_type == "mz_rt_reference":
        return f"mzrt-{uuid.uuid4().hex[:32]}"
    elif entity_type == "association":
        return f"assoc-{uuid.uuid4().hex[:32]}"
    elif entity_type == "rt_alignment":
        return f"rta-{uuid.uuid4().hex[:32]}"
    elif entity_type == "analysis":
        return f"tga-{uuid.uuid4().hex[:32]}"
    else:
        raise ValueError(f"Unknown entity type: {entity_type}")

def get_lcmsruns_from_db(project_db_path: str, file_types: List[str] = None) -> List[str]:
    """Get experimental files from project database."""
    
    if file_types is None:
        file_types = ['experimental', 'istd', 'exctrl']
    
    with get_db_connection(project_db_path) as conn:

        # Get experimental files (excluding QC)
        placeholders = ','.join(['?' for _ in file_types])
        result = conn.execute(f"""
            SELECT file_path 
            FROM lcmsruns 
            WHERE file_type IN ({placeholders})
            ORDER BY filename
        """, file_types).fetchall()
        
    file_paths = [row[0] for row in result]
    logger.info(f"Retrieved {len(file_paths)} files in database")
    
    return file_paths

def get_atlas_compounds_from_db(db_path: str, atlas_uid: str, main_db_path: str = None) -> pd.DataFrame:
    """
    Get atlas compounds from database - wrapper around get_atlas_compounds_table.
    """
    return get_atlas_compounds_table(db_path, atlas_uid, main_db_path)

# def get_atlas_compounds_with_metadata(project_db_path: str, main_db_path: str, atlas_uid: str) -> pd.DataFrame:
#     """
#     Get atlas compounds from project database and enrich with metadata from main database.
#     This is specifically for RT-corrected atlases in project databases.
#     Uses external database attachment for compound metadata.
#     """
#     # Use the updated get_atlas_compounds_table with external attachment
#     enriched_df = get_atlas_compounds_table(project_db_path, atlas_uid, main_db_path)
    
#     if enriched_df.empty:
#         return pd.DataFrame()
    
#     # Add label column for compatibility with feature_tools
#     enriched_df['label'] = enriched_df['compound_name']
    
#     logger.info(f"Retrieved {len(enriched_df)} compounds with metadata using external database attachment")
    
#     return enriched_df

# def save_targeted_analysis_from_analysis_project(
#     analysis_project,
#     project_name: str,
#     atlas_uid: str
# ) -> str:
#     """
#     Save targeted analysis results using AnalysisProject class.
#     This replaces the old deposit_targeted_analysis_from_plot_data function.
#     """
#     logger.info(f"Saving targeted analysis results for project '{project_name}' using class-based approach...")
    
#     # Use the class method to save to database
#     analysis_uid = analysis_project.save_to_database(project_name, atlas_uid)
    
#     # Validate the saved data
#     validated = validate_targeted_analysis_data(analysis_project.project_db_path, analysis_uid)
#     if not validated:
#         logger.warning(f"Validation failed for targeted analysis entry {analysis_uid}")
#         return None
    
#     logger.info(f"Successfully saved targeted analysis with UID: {analysis_uid}")
#     return analysis_uid

# def generate_comprehensive_targeted_analysis_report(project_db_path: str, config: Dict, 
#                                                    analysis_uid: str, atlas_dataframe: pd.DataFrame,
#                                                    post_analysis_atlas_uid: str,
#                                                    output_path: str = None, include_missing_compounds: bool = False) -> pd.DataFrame:
#     """
#     Generate a comprehensive targeted analysis report matching the specified Excel format.
#     Updated to work with consistent per-file MS2 data structure.
#     """
#     main_db_path = config["ENV"]["PATHS"]["main_database"]
    
#     # Get the base targeted analysis summary
#     base_summary = generate_targeted_analysis_summary(project_db_path, config, analysis_uid)
    
#     if base_summary.empty:
#         logger.warning(f"No targeted analysis results found for analysis_uid {analysis_uid}")
#         if atlas_dataframe is not None and include_missing_compounds:
#             logger.info("Creating empty rows for all atlas compounds...")
#             base_summary = _create_empty_summary_from_atlas(atlas_dataframe, analysis_uid, config)
#         else:
#             return pd.DataFrame()
#     elif atlas_dataframe is not None and include_missing_compounds:
#         # Add missing compounds from atlas as empty rows
#         base_summary = _add_missing_compounds_to_summary(base_summary, atlas_dataframe, analysis_uid, config)
    
#     # Connect to both databases
#     with get_db_connection(project_db_path) as conn_proj:
#         with get_db_connection(main_db_path) as conn_main:
            
#             report_rows = []
            
#             logger.info(f"Generating comprehensive report for {len(base_summary)} compounds...")
            
#             for idx, row in tqdm(base_summary.iterrows(), total=len(base_summary), desc="Processing compounds"):
#                 try:
#                     # Basic compound information
#                     compound_uid = row['compound_uid']
#                     inchi_key = row['inchi_key']
#                     compound_name = row['compound_name']
                    
#                     # Calculate stats using simplified methods
#                     msms_quality = _calculate_msms_quality_score_from_notes(row)
#                     mz_quality = _calculate_mz_quality_score(row)
#                     rt_quality = _calculate_rt_quality_score(row)
#                     total_score = msms_quality + mz_quality + rt_quality
#                     msi_level = _determine_msi_level(msms_quality, mz_quality, rt_quality)
                    
#                     # Find isomers using simplified approach
#                     isomer_info = _find_overlapping_compounds_simple(conn_main, row, base_summary)
                    
#                     # Build the report row with consistent field names
#                     report_row = {
#                         'index': idx,
#                         'identified_metabolite': compound_name,
#                         'label': compound_name,
#                         'isomer_compound': isomer_info['compound_names'],
#                         'isomer_inchi_keys': isomer_info['inchi_keys'],
#                         'formula': row.get('formula', ''),
#                         'polarity': row.get('polarity', ''),
#                         'mono_isotopic_molecular_weight': row.get('mono_isotopic_molecular_weight', ''),
#                         'inchi_key': inchi_key,
#                         'adduct': row.get('adduct', ''),
#                         'msms_quality': msms_quality,
#                         'mz_quality': mz_quality,
#                         'rt_quality': rt_quality,
#                         'total_score': total_score,
#                         'msi_level': msi_level,
#                         'ms1_notes': row.get('ms1_notes', 'keep'),
#                         'ms2_notes': row.get('ms2_notes', 'no selection'),
#                         'analyst_notes': row.get('analyst_notes', ''),
#                         'identification_notes': row.get('identification_notes', ''),
#                         'max_intensity': row.get('best_eic_intensity', ''),
#                         'max_intensity_file': row.get('best_eic_file', ''),
#                         'max_intensity_rt': row.get('best_eic_rt', ''),
#                         'best_msms_file': row.get('best_ms2_file', ''),
#                         'best_msms_rt': row.get('best_ms2_rt_peak', ''),
#                         'best_msms_num_matching_ions': row.get('best_ms2_num_matches', ''),
#                         'best_msms_matching_ions': row.get('best_ms2_matched_fragments', ''),
#                         'best_msms_score': row.get('best_ms2_score', ''),
#                         'mz_theoretical': row.get('pre_mz', ''),
#                         'mz_measured': row.get('best_eic_mz', ''),
#                         'mz_error': row.get('best_eic_ppm_error', ''),
#                         'rt_peak_theoretical': row.get('pre_rt_peak', ''),
#                         'rt_peak_measured': row.get('post_rt_peak', ''),
#                         'rt_min_measured': row.get('post_rt_min', ''),
#                         'rt_max_measured': row.get('post_rt_max', ''),
#                         'rt_error': row.get('best_eic_rt_error', '')
#                     }
                    
#                     report_rows.append(report_row)
                    
#                 except Exception as e:
#                     logger.error(f"Error processing compound {compound_name} ({inchi_key}): {e}")
#                     continue
    
#     # Create DataFrame and sort
#     report_df = pd.DataFrame(report_rows)
    
#     if not report_df.empty:
#         # Convert rt_peak_theoretical to numeric for proper sorting
#         report_df['rt_peak_theoretical_num'] = pd.to_numeric(report_df['rt_peak_theoretical'], errors='coerce')
#         report_df = report_df.sort_values('rt_peak_theoretical_num', ascending=True, na_position='last')
#         report_df = report_df.drop(columns=['rt_peak_theoretical_num'])
#         report_df = report_df.reset_index(drop=True)
#         report_df['index'] = range(len(report_df))
    
#     # Save to Excel if path provided
#     if output_path is not None and not report_df.empty:
#         try:
#             _save_report_with_grouped_headers(report_df, output_path, post_analysis_atlas_uid)
#             logger.info(f"Report saved to {output_path}")
#         except Exception as e:
#             logger.error(f"Error saving Excel file: {e}")
    
#     return report_df

def _find_overlapping_compounds_simple(conn_main, current_row: pd.Series, all_compounds: pd.DataFrame) -> Dict[str, str]:
    """
    Find overlapping compounds using simplified approach.
    Uses the isomers field if available, otherwise falls back to database lookup.
    """
    # First try to use the isomers field if it exists and is populated
    isomers_json = current_row.get('isomers', None)
    if isomers_json and isomers_json != 'null':
        try:
            import json
            isomers = json.loads(isomers_json) if isinstance(isomers_json, str) else isomers_json
            if isinstance(isomers, list) and isomers:
                compound_names = []
                inchi_keys = []
                for iso in isomers:
                    name = iso.get('compound_name', 'Unknown')
                    inchi = iso.get('inchi_key', '')
                    if inchi and inchi != current_row.get('inchi_key', ''):
                        compound_names.append(name)
                        inchi_keys.append(inchi)
                return {
                    'compound_names': '; '.join(compound_names),
                    'inchi_keys': '; '.join(inchi_keys),
                }
        except (json.JSONDecodeError, TypeError):
            pass
    
    # Fallback to empty if no isomers data
    return {
        'compound_names': '',
        'inchi_keys': '',
    }

def batch_save_compounds_and_references(
    compounds_data: List[Dict],
    references_data: List[Dict],
    db_path: str
) -> Tuple[int, int]:
    """
    Schema-compliant batch save for compounds and references from raw input data.
    
    Schema Logic:
    1. For each compound_data: check if inchi_key exists in database
       - If exists: get existing compound_uid, skip compound creation
       - If not exists: create new compound with new compound_uid
    2. For each reference_data: find corresponding compound_uid, then check if identical reference exists
       - If identical: skip reference creation
       - If different: create new reference entry
    
    Args:
        compounds_data: List of compound dictionaries
        references_data: List of reference dictionaries
        db_path: str

    Returns:
        Tuple of (compounds_created, references_created)
    """
    compounds_created = 0
    references_created = 0
    references_skipped_identical = 0
    compounds_skipped_existing = 0
    
    with get_db_connection(db_path) as conn:
        # Step 1: Get ALL existing compounds (inchi_key -> compound_uid mapping)
        existing_compounds = {}  # inchi_key -> compound_uid
        existing_result = conn.execute("SELECT inchi_key, compound_uid FROM compounds").fetchall()
        for row in existing_result:
            existing_compounds[row[0]] = row[1]
        
        logger.info(f"Found {len(existing_compounds)} existing compounds in database")
        
        # Step 2: Process compound data and build inchi_key -> compound_uid mapping
        compound_records = []
        inchi_to_compound_uid = {}  # This will contain both existing and new compound mappings
        
        for compound_data in compounds_data:
            inchi_key = compound_data.get('inchi_key')
            if not inchi_key:
                logger.warning("Skipping compound with missing inchi_key")
                continue
                
            if inchi_key in existing_compounds:
                # Compound already exists - use existing UID
                inchi_to_compound_uid[inchi_key] = existing_compounds[inchi_key]
                compounds_skipped_existing += 1
                logger.debug(f"Compound already exists: {inchi_key} -> {existing_compounds[inchi_key]}")
            else:
                # New compound - generate UID and prepare for creation
                new_compound_uid = _generate_uid("compound")
                compound_record = _prepare_compound_record_from_dict({
                    **compound_data,  # Raw data from input
                    'compound_uid': new_compound_uid  # Add the generated UID
                })
                
                if compound_record:
                    compound_records.append(compound_record)
                    inchi_to_compound_uid[inchi_key] = new_compound_uid
                    existing_compounds[inchi_key] = new_compound_uid  # Prevent duplicates within batch
                    compounds_created += 1
                    logger.debug(f"New compound prepared: {inchi_key} -> {new_compound_uid}")
        
        # Step 3: Batch insert new compounds
        if compound_records:
            logger.info(f"Creating {len(compound_records)} new compounds...")
            conn.executemany("""
                INSERT INTO compounds VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, compound_records)
        
        # Step 4: Process reference data - match to compounds and check for duplicates
        if references_data:
            logger.info(f"Processing {len(references_data)} reference entries...")
            reference_records = []
            
            for reference_data in references_data:
                # Find the compound_uid for this reference using inchi_key
                ref_inchi_key = reference_data.get('inchi_key')
                if not ref_inchi_key:
                    logger.warning("Skipping reference with missing inchi_key")
                    continue
                
                compound_uid = inchi_to_compound_uid.get(ref_inchi_key)
                if not compound_uid:
                    logger.warning(f"No compound found for reference with inchi_key: {ref_inchi_key}")
                    continue
                
                # Prepare reference data with compound_uid
                reference_with_uid = {
                    **reference_data,
                    'compound_uid': compound_uid
                }
                
                # Check if IDENTICAL reference already exists
                identical_ref_exists = _check_identical_reference_exists(conn, reference_with_uid)
                
                if identical_ref_exists:
                    references_skipped_identical += 1
                    logger.debug(f"Identical reference already exists for compound {compound_uid}")
                    continue
                
                # Create new reference entry
                reference_record = _prepare_reference_record_from_dict(reference_with_uid)
                if reference_record:
                    reference_records.append(reference_record)
                    references_created += 1
                    logger.debug(f"New reference prepared for compound {compound_uid}")
            
            # Step 5: Batch insert new references
            if reference_records:
                logger.info(f"Creating {len(reference_records)} new references...")
                conn.executemany("""
                    INSERT INTO mz_rt_references VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, reference_records)
    
    # Log summary
    logger.info("Batch save completed:")
    logger.info(f"  Compounds created: {compounds_created}")
    logger.info(f"  Compounds skipped (already exist): {compounds_skipped_existing}")
    logger.info(f"  References created: {references_created}")
    logger.info(f"  References skipped (identical data): {references_skipped_identical}")
    
    return compounds_created, references_created


def _check_identical_reference_exists(conn, reference_data: Dict) -> bool:
    """
    Check if an IDENTICAL reference already exists in the database.
    This implements the schema rule: only create new references if data is different.
    
    Args:
        conn: Database connection
        reference_data: Dictionary with reference data
        
    Returns:
        True if identical reference exists, False otherwise
    """
    try:
        compound_uid = reference_data.get('compound_uid')
        rt_peak = float(reference_data.get('rt_peak', 0.0))
        rt_min = float(reference_data.get('rt_min', 0.0))
        rt_max = float(reference_data.get('rt_max', 0.0))
        mz = float(reference_data.get('mz', 0.0))
        mz_tolerance = float(reference_data.get('mz_tolerance', 5.0))
        adduct = str(reference_data.get('adduct', ''))
        chromatography = str(reference_data.get('chromatography', ''))
        polarity = str(reference_data.get('polarity', ''))
        
        # Check for IDENTICAL reference (using small tolerance for floating point comparison)
        existing_ref = conn.execute("""
            SELECT mz_rt_reference_uid 
            FROM mz_rt_references 
            WHERE compound_uid = ? 
            AND chromatography = ? 
            AND polarity = ? 
            AND adduct = ?
            AND ABS(rt_peak - ?) < 0.0001
            AND ABS(rt_min - ?) < 0.0001  
            AND ABS(rt_max - ?) < 0.0001
            AND ABS(mz - ?) < 0.0001
            AND ABS(mz_tolerance - ?) < 0.0001
        """, [
            compound_uid, chromatography, polarity, adduct,
            rt_peak, rt_min, rt_max, mz, mz_tolerance
        ]).fetchone()
        
        return existing_ref is not None
        
    except (ValueError, TypeError) as e:
        logger.warning(f"Error checking reference data: {e}")
        return False

def _create_database_tables(conn, db_type: str = "main"):
    """
    Create database tables based on database type.
    
    Args:
        conn: Database connection
        db_type: Either "main" or "project"
    """
    if db_type == "main":
        conn.execute("""
            CREATE TABLE IF NOT EXISTS compounds (
                compound_uid TEXT PRIMARY KEY,
                name TEXT,
                inchi_key TEXT,
                inchi TEXT,
                smiles TEXT,
                formula TEXT,
                compound_classes TEXT,
                compound_pathways TEXT,
                compound_tags TEXT,
                mono_isotopic_molecular_weight REAL,
                iupac_name TEXT,
                pubchem_cid TEXT,
                cas_number TEXT,
                synonyms TEXT,
                created_by TEXT,
                created_date TEXT
            )
        """)
        
        conn.execute("""
            CREATE TABLE IF NOT EXISTS mz_rt_references (
                mz_rt_reference_uid TEXT PRIMARY KEY,
                compound_uid TEXT,
                rt_peak REAL,
                rt_min REAL,
                rt_max REAL,
                mz REAL,
                mz_tolerance REAL,
                adduct TEXT,
                chromatography TEXT,
                polarity TEXT,
                confidence TEXT,
                source TEXT,
                created_by TEXT,
                created_date TEXT,
                FOREIGN KEY (compound_uid) REFERENCES compounds (compound_uid)
            )
        """)
        
        conn.execute("""
            CREATE TABLE IF NOT EXISTS atlases (
                atlas_uid TEXT PRIMARY KEY,
                atlas_name TEXT,
                atlas_description TEXT,
                chromatography TEXT,
                polarity TEXT,
                atlas_type TEXT,
                created_by TEXT,
                last_modified TEXT
            )
        """)
        
        conn.execute("""
            CREATE TABLE IF NOT EXISTS atlas_compound_associations (
                association_uid TEXT PRIMARY KEY,
                atlas_uid TEXT,
                compound_uid TEXT,
                mz_rt_reference_uid TEXT,
                association_order INTEGER,
                created_by TEXT,
                created_date TEXT,
                FOREIGN KEY (atlas_uid) REFERENCES atlases (atlas_uid),
                FOREIGN KEY (compound_uid) REFERENCES compounds (compound_uid),
                FOREIGN KEY (mz_rt_reference_uid) REFERENCES mz_rt_references (mz_rt_reference_uid)
            )
        """)
        
    elif db_type == "project":
        conn.execute("""
            CREATE TABLE IF NOT EXISTS lcmsruns (
                file_path TEXT PRIMARY KEY,
                filename TEXT,
                file_type TEXT,
                chromatography TEXT,
                polarity TEXT,
                created_by TEXT,
                created_date TEXT
            )
        """)
        
        conn.execute("""
            CREATE TABLE IF NOT EXISTS atlases (
                atlas_uid TEXT PRIMARY KEY,
                atlas_name TEXT,
                atlas_description TEXT,
                chromatography TEXT,
                polarity TEXT,
                atlas_type TEXT,
                source_atlas_uid TEXT,
                rt_alignment_number INTEGER,
                analysis_number INTEGER,
                created_by TEXT,
                last_modified TEXT
            )
        """)
        
        conn.execute("""
            CREATE TABLE IF NOT EXISTS compounds (
                compound_uid TEXT PRIMARY KEY,
                name TEXT,
                inchi_key TEXT,
                inchi TEXT,
                smiles TEXT,
                formula TEXT,
                compound_classes TEXT,
                compound_pathways TEXT,
                compound_tags TEXT,
                mono_isotopic_molecular_weight REAL,
                iupac_name TEXT,
                pubchem_cid TEXT,
                cas_number TEXT,
                synonyms TEXT,
                created_by TEXT,
                created_date TEXT
            )
        """)
        
        conn.execute("""
            CREATE TABLE IF NOT EXISTS mz_rt_experimental (
                mz_rt_experimental_uid TEXT PRIMARY KEY,
                compound_uid TEXT,
                rt_alignment_number INTEGER,
                analysis_number INTEGER,
                rt_peak REAL,
                rt_min REAL,
                rt_max REAL,
                ms1_notes TEXT,
                ms2_notes TEXT,
                mz REAL,
                mz_tolerance REAL,
                adduct TEXT,
                chromatography TEXT,
                polarity TEXT,
                rt_correction_applied BOOLEAN,
                rt_shift REAL,
                source_mz_rt_reference_uid TEXT,
                created_by TEXT,
                created_date TEXT
            )
        """)
        
        conn.execute("""
            CREATE TABLE IF NOT EXISTS atlas_compound_associations (
                association_uid TEXT PRIMARY KEY,
                atlas_uid TEXT,
                compound_uid TEXT,
                mz_rt_reference_uid TEXT,
                association_order INTEGER,
                created_by TEXT,
                created_date TEXT,
                FOREIGN KEY (atlas_uid) REFERENCES atlases (atlas_uid)
            )
        """)
        
        conn.execute("""
            CREATE TABLE IF NOT EXISTS rt_alignment (
                rt_alignment_uid TEXT PRIMARY KEY,
                project_name TEXT,
                rt_alignment_number INTEGER,
                qc_atlas_uid TEXT,
                model_type TEXT,
                polynomial_degree INTEGER,
                r_squared REAL,
                rmse REAL,
                coefficients TEXT,
                equation TEXT,
                num_qc_files INTEGER,
                num_compounds INTEGER,
                created_by TEXT,
                created_date TEXT,
                metadata TEXT
            )
        """)
        
        conn.execute("""
            CREATE TABLE IF NOT EXISTS targeted_analysis (
                analysis_uid TEXT,
                project_name TEXT,
                rt_alignment_number INTEGER,
                analysis_number INTEGER,
                atlas_uid TEXT,
                atlas_type TEXT,
                chromatography_polarity TEXT,
                compound_uid TEXT,
                inchi_key TEXT,
                compound_name TEXT,
                pre_rt_peak REAL,
                pre_rt_min REAL,
                pre_rt_max REAL,
                pre_mz REAL,
                mz_tolerance REAL,
                adduct TEXT,
                isomers TEXT,
                post_rt_peak REAL,
                post_rt_min REAL,
                post_rt_max REAL,
                is_rt_modified BOOLEAN,
                best_eic_file TEXT,
                best_eic_rt REAL,
                best_eic_mz REAL,
                best_eic_intensity REAL,
                best_eic_ppm_error REAL,
                best_eic_rt_error REAL,
                avg_eic_rt REAL,
                avg_eic_intensity REAL,
                avg_eic_mz REAL,
                best_ms2_file TEXT,
                best_ms2_database TEXT,
                best_ms2_ref_id TEXT,
                best_ms2_rt_peak REAL,
                best_ms2_intensity_peak REAL,
                best_ms2_mz_peak REAL,
                best_ms2_score REAL,
                best_ms2_num_matches INTEGER,
                best_ms2_ref_frags INTEGER,
                best_ms2_data_frags INTEGER,
                best_ms2_matched_fragments TEXT,
                avg_ms2_score REAL,
                total_files_detected INTEGER,
                ms2_files_with_data INTEGER,
                ms2_best_score REAL,
                ms2_best_database TEXT,
                ms2_total_matches INTEGER,
                ms1_notes TEXT,
                ms2_notes TEXT,
                analyst_notes TEXT,
                identification_notes TEXT,
                curation_status TEXT,
                analyst TEXT,
                analysis_timestamp TEXT,
                PRIMARY KEY (analysis_uid, compound_uid, rt_alignment_number, analysis_number)
            )
        """)

# def create_atlas_from_compounds(atlas_compounds_df: pd.DataFrame, atlas_name: str, 
#                                atlas_description: str, atlas_type: str,
#                                chromatography: str, polarity: str,
#                                config: Dict) -> Tuple[str, str]:
#     """
#     Create a new atlas from a DataFrame of compounds.
    
#     Args:
#         atlas_compounds_df: DataFrame with compound and RT/MZ reference data
#         atlas_name: Name for the new atlas
#         atlas_description: Description for the new atlas
#         config: Configuration dictionary
    
#     Returns:
#         Tuple of (atlas_uid, atlas_name)
#     """
#     db_path = config["ENV"]["PATHS"]["main_database"]
    
#     # Generate atlas UID
#     atlas_uid = _generate_uid("atlas", decorator=atlas_type.lower())
    
#     prov = ldt.get_provenance()
    
#     with get_db_connection(db_path) as conn:
#         # Create atlas
#         conn.execute("""
#             INSERT INTO atlases VALUES (?, ?, ?, ?, ?, ?, ?)
#         """, (
#             atlas_uid,
#             atlas_name,
#             atlas_description,
#             chromatography,
#             polarity,
#             atlas_type,
#             prov["analyst"],
#             prov["timestamp"]
#         ))
        
#         # Get existing compounds
#         existing_compounds = {}
#         existing_compound_result = conn.execute("SELECT inchi_key, compound_uid FROM compounds").fetchall()
#         for row in existing_compound_result:
#             existing_compounds[row[0]] = row[1]
#         logger.info(f"Found {len(existing_compounds)} existing compounds in database")
        
#         # Process compounds and create/find RT/MZ references
#         association_order = 0
#         compounds_processed = 0
#         compounds_skipped = 0
#         references_created = 0
#         references_reused = 0
        
#         for _, row in atlas_compounds_df.iterrows():
#             inchi_key = row.get('inchi_key', '')
#             if not inchi_key:
#                 continue
            
#             # Check if compound exists in database
#             compound_uid = existing_compounds.get(inchi_key)
#             if not compound_uid:
#                 logger.warning(f"Compound {inchi_key} not found in database, skipping")
#                 compounds_skipped += 1
#                 continue
            
#             # Extract RT/MZ data from input
#             rt_peak = row.get('rt_peak')
#             rt_min = row.get('rt_min')
#             rt_max = row.get('rt_max')
#             mz = row.get('mz')
#             adduct = row.get('adduct', '')
#             mz_tolerance = row.get('mz_tolerance', 5.0)
            
#             # Validate that we have the required RT/MZ data
#             if pd.isna(rt_peak) or pd.isna(mz):
#                 logger.warning(f"Compound {inchi_key} missing RT or MZ data, skipping")
#                 compounds_skipped += 1
#                 continue
            
#             # Check for existing reference with identical data
#             existing_ref_query = """
#                 SELECT mz_rt_reference_uid 
#                 FROM mz_rt_references 
#                 WHERE compound_uid = ? 
#                 AND chromatography = ? 
#                 AND polarity = ? 
#                 AND adduct = ?
#                 AND ABS(rt_peak - ?) < 0.001 
#                 AND ABS(mz - ?) < 0.001
#                 AND ABS(mz_tolerance - ?) < 0.001
#             """
            
#             existing_ref = conn.execute(existing_ref_query, [
#                 compound_uid, chromatography, polarity, adduct,
#                 float(rt_peak), float(mz), float(mz_tolerance)
#             ]).fetchone()
            
#             if existing_ref:
#                 # Use existing identical reference
#                 mz_rt_reference_uid = existing_ref[0]
#                 references_reused += 1
#                 logger.debug(f"Reusing existing RT/MZ reference for compound {inchi_key}")
#             else:
#                 # Create new RT/MZ reference
#                 mz_rt_reference_uid = _generate_uid("mz_rt_reference")
                
#                 # Prepare RT bounds
#                 if pd.isna(rt_min):
#                     rt_min = float(rt_peak) - 0.5
#                 if pd.isna(rt_max):
#                     rt_max = float(rt_peak) + 0.5
                
#                 conn.execute("""
#                     INSERT INTO mz_rt_references VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
#                 """, (
#                     mz_rt_reference_uid,
#                     compound_uid,
#                     float(rt_peak),
#                     float(rt_min),
#                     float(rt_max),
#                     float(mz),
#                     float(mz_tolerance),
#                     adduct,
#                     chromatography,
#                     polarity,
#                     row.get('confidence', 'Unknown'),
#                     'atlas_creation',  # source
#                     prov["analyst"],
#                     prov["timestamp"]
#                 ))
                
#                 references_created += 1
#                 logger.debug(f"Created new RT/MZ reference for compound {inchi_key}")
            
#             # Create atlas-compound association
#             assoc_uid = _generate_uid("association")
#             conn.execute("""
#                 INSERT INTO atlas_compound_associations VALUES (?, ?, ?, ?, ?, ?, ?)
#             """, (
#                 assoc_uid,
#                 atlas_uid,
#                 compound_uid,
#                 mz_rt_reference_uid,
#                 association_order,
#                 prov["analyst"],
#                 prov["timestamp"]
#             ))
            
#             association_order += 1
#             compounds_processed += 1
    
    
#     logger.info(f"Created atlas '{atlas_name}' with UID: {atlas_uid}")
#     logger.info(f"  Processed {compounds_processed} compounds successfully")
#     logger.info(f"  Skipped {compounds_skipped} compounds (missing from database or invalid data)")
#     logger.info(f"  Created {references_created} new RT/MZ references")
#     logger.info(f"  Reused {references_reused} existing RT/MZ references")
#     logger.info(f"  Added {association_order} compound associations")
    
#     return atlas_uid, atlas_name

def verify_compounds_exist_in_db(compound_uids: list, db_path: str) -> pd.DataFrame:
    """
    Verify that all compound_uids exist in the database at db_path.
    Returns a DataFrame of compounds that exist in the database.
    Logs warnings for any missing compound_uids.
    """
    if not compound_uids:
        logger.warning("No compound_uids provided for verification.")
        return pd.DataFrame()

    with get_db_connection(db_path) as conn:
        placeholders = ','.join(['?'] * len(compound_uids))
        query = f"SELECT compound_uid FROM compounds WHERE compound_uid IN ({placeholders})"
        existing = conn.execute(query, compound_uids).fetchall()
        existing_uids = {row[0] for row in existing}

    missing_uids = set(compound_uids) - existing_uids
    if missing_uids:
        for uid in missing_uids:
            logger.warning(f"Compound {uid} not found in database {db_path}")
        return False
    return True

def get_compound_uids_by_inchi_keys(db_path: str, inchi_keys: list) -> dict:
    """Return {inchi_key: compound_uid} for all inchi_keys found in the database."""
    with get_db_connection(db_path) as conn:
        if not inchi_keys:
            return {}
        placeholders = ','.join(['?'] * len(inchi_keys))
        rows = conn.execute(
            f"SELECT inchi_key, compound_uid FROM compounds WHERE inchi_key IN ({placeholders})",
            inchi_keys
        ).fetchall()
        return {row[0]: row[1] for row in rows}

def get_or_create_mz_rt_reference_uid(
    db_path: str,
    compound_uid: str,
    chromatography: str,
    polarity: str,
    adduct: str,
    rt_peak: float,
    mz: float,
    mz_tolerance: float
) -> (str, bool):
    """
    Return (mz_rt_reference_uid, reused_flag). If not found, generate a new UID (do not insert).
    """
    with get_db_connection(db_path) as conn:
        existing = conn.execute("""
            SELECT mz_rt_reference_uid FROM mz_rt_references
            WHERE compound_uid = ?
            AND chromatography = ? AND polarity = ? AND adduct = ?
            AND ABS(rt_peak - ?) < 0.0001 AND ABS(mz - ?) < 0.0001
            AND ABS(mz_tolerance - ?) < 0.0001
        """, [
            compound_uid, chromatography, polarity, adduct,
            rt_peak, mz, mz_tolerance
        ]).fetchone()
        if existing:
            return existing[0], True
        else:
            return _generate_uid("mz_rt_reference"), False

def save_atlas_to_database(atlas_obj: Atlas, db_path: str, db_type: str = "main") -> None:
    """
    Save an Atlas object to the database.
    Only creates new mz_rt_references entries for references that don't already exist.
    
    Args:
        atlas_obj: Atlas object to save
    """
    # Verify all compounds exist in database
    if not verify_compounds_exist_in_db([comp.compound_uid for comp in atlas_obj.compound_references.values()], db_path):
        raise ValueError(f"Some compounds in atlas {atlas_obj.atlas_uid} don't exist in database")

    prov = ldt.get_provenance()
    with dbi.get_db_connection(db_path) as conn:
        if db_type == "main":
            # Create atlas entry
            conn.execute("""
                INSERT INTO atlases VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                atlas_obj.atlas_uid,
                atlas_obj.atlas_name,
                atlas_obj.atlas_description,
                atlas_obj.chromatography,
                atlas_obj.polarity,
                "REFERENCE",  # atlas_type
                prov["analyst"],
                prov["timestamp"]
            ))
            
            # Process each CompoundReference
            association_order = 0
            references_created = 0
            references_reused = 0
            for inchi_key, compound_ref in atlas_obj.compound_references.items():
                # Check if this reference already exists in database
                existing_check = conn.execute("""
                    SELECT mz_rt_reference_uid FROM mz_rt_references 
                    WHERE mz_rt_reference_uid = ?
                """, [compound_ref.mz_rt_reference_uid]).fetchone()
                
                mz_rt_reference_uid = compound_ref.mz_rt_reference_uid
                
                # Create new reference if it doesn't exist
                if not existing_check:
                    conn.execute("""
                        INSERT INTO mz_rt_references VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        mz_rt_reference_uid,
                        compound_ref.compound_uid,
                        compound_ref.rt_peak,
                        compound_ref.rt_min,
                        compound_ref.rt_max,
                        compound_ref.mz,
                        compound_ref.mz_tolerance,
                        compound_ref.adduct,
                        compound_ref.chromatography,
                        compound_ref.polarity,
                        compound_ref.confidence,
                        'atlas_creation',  # source
                        prov["analyst"],
                        prov["timestamp"]
                    ))
                    references_created += 1
                else:
                    references_reused += 1
                
                # Create atlas-compound association
                assoc_uid = dbi._generate_uid("association")
                conn.execute("""
                    INSERT INTO atlas_compound_associations VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    assoc_uid,
                    atlas_obj.atlas_uid,
                    compound_ref.compound_uid,
                    mz_rt_reference_uid,
                    association_order,
                    prov["analyst"],
                    prov["timestamp"]
                ))
                
                association_order += 1
        
            logger.info(f"Saved atlas {atlas_obj.atlas_name} to database with UID: {atlas_obj.atlas_uid}")
            logger.info(f"  References created: {references_created}")
            logger.info(f"  References reused: {references_reused}")
            logger.info(f"  Total associations: {association_order}")

        else:
            # Generate new atlas UID for RT-corrected version
            corrected_atlas_uid = _generate_uid("rt_atlas")
            corrected_atlas_name = f"{target_atlas_info['atlas_name']} (RT Corrected)"
            corrected_atlas_description = f"RT-corrected version of {target_atlas_info['atlas_name']} using polynomial model (R²={best_model['r2']:.4f})"

            # Create new atlas entry
            conn.execute("""
                INSERT INTO atlases VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                corrected_atlas_uid,
                corrected_atlas_name,
                corrected_atlas_description,
                target_atlas_info['chromatography'],
                target_atlas_info['polarity'],
                target_atlas_info['atlas_type'],
                target_atlas_info['atlas_uid'],
                rt_alignment_number,
                analysis_number,
                prov["analyst"],
                prov["timestamp"]
            ))

            association_order = 0
            for _, row in corr_compounds_df.iterrows():
                compound_uid = row['compound_uid']
                corrected_rt_peak = row.get('rt_peak')
                corrected_rt_min = row.get('rt_min')
                corrected_rt_max = row.get('rt_max')
                rt_shift = row.get('rt_shift')
                exp_uid = _generate_uid("mz_rt_experimental")
                assoc_uid = _generate_uid("association")

                conn.execute("""
                    INSERT INTO mz_rt_experimental VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    exp_uid,
                    compound_uid,
                    rt_alignment_number,
                    analysis_number,
                    corrected_rt_peak,
                    corrected_rt_min,
                    corrected_rt_max,
                    '',  # ms1_notes
                    '',  # ms2_notes
                    row.get('mz'),
                    row.get('mz_tolerance', 5.0),
                    row.get('adduct', ''),
                    target_atlas_info['chromatography'],
                    target_atlas_info['polarity'],
                    True,  # rt_correction_applied
                    rt_shift,
                    row.get('mz_rt_reference_uid'),  # source_mz_rt_reference_uid
                    prov["analyst"],
                    prov["timestamp"]
                ))
                
                # Create atlas-compound association
                conn.execute("""
                    INSERT INTO atlas_compound_associations VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    assoc_uid,
                    corrected_atlas_uid,
                    compound_uid,
                    exp_uid,
                    association_order,
                    prov["analyst"],
                    prov["timestamp"]
                ))

        association_order += 1

        logger.info(f"Created RT-corrected atlas: {corrected_atlas_uid} name: {corrected_atlas_name}")

def save_rt_corrected_atlas_to_db(
    project_db_path: str,
    target_atlas_info: pd.DataFrame,
    best_model: dict,
    corr_compounds_df: pd.DataFrame,
    rt_alignment_number: int,
    analysis_number: int
) -> Tuple[str, str]:
    """
    Create RT-corrected atlas in project database and apply RT correction to compounds.
    
    Args:
        project_db_path: Path to project database
        target_atlas_info: DataFrame containing target atlas metadata
        best_model: RT correction model dictionary
        corr_compounds_df: DataFrame of RT corrected compounds
    
    Returns:
        Tuple of (corrected_atlas_uid, corrected_atlas_name)
    """
    logger.info("Creating RT-corrected atlas in project database...")

    prov = ldt.get_provenance()

    # Generate new atlas UID for RT-corrected version
    corrected_atlas_uid = _generate_uid("rt_atlas")
    corrected_atlas_name = f"{target_atlas_info['atlas_name']} (RT Corrected)"
    corrected_atlas_description = f"RT-corrected version of {target_atlas_info['atlas_name']} using polynomial model (R²={best_model['r2']:.4f})"

    with get_db_connection(project_db_path) as conn:
        
        # Create new atlas entry
        conn.execute("""
            INSERT INTO atlases VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            corrected_atlas_uid,
            corrected_atlas_name,
            corrected_atlas_description,
            target_atlas_info['chromatography'],
            target_atlas_info['polarity'],
            target_atlas_info['atlas_type'],
            target_atlas_info['atlas_uid'],
            rt_alignment_number,
            analysis_number,
            prov["analyst"],
            prov["timestamp"]
        ))

        association_order = 0
        for _, row in corr_compounds_df.iterrows():
            compound_uid = row['compound_uid']
            corrected_rt_peak = row.get('rt_peak')
            corrected_rt_min = row.get('rt_min')
            corrected_rt_max = row.get('rt_max')
            rt_shift = row.get('rt_shift')
            exp_uid = _generate_uid("mz_rt_experimental")
            assoc_uid = _generate_uid("association")

            conn.execute("""
                INSERT INTO mz_rt_experimental VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                exp_uid,
                compound_uid,
                rt_alignment_number,
                analysis_number,
                corrected_rt_peak,
                corrected_rt_min,
                corrected_rt_max,
                '',  # ms1_notes
                '',  # ms2_notes
                row.get('mz'),
                row.get('mz_tolerance', 5.0),
                row.get('adduct', ''),
                target_atlas_info['chromatography'],
                target_atlas_info['polarity'],
                True,  # rt_correction_applied
                rt_shift,
                row.get('mz_rt_reference_uid'),  # source_mz_rt_reference_uid
                prov["analyst"],
                prov["timestamp"]
            ))
            
            # Create atlas-compound association
            conn.execute("""
                INSERT INTO atlas_compound_associations VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                assoc_uid,
                corrected_atlas_uid,
                compound_uid,
                exp_uid,
                association_order,
                prov["analyst"],
                prov["timestamp"]
            ))

    association_order += 1

    logger.info(f"Created RT-corrected atlas: {corrected_atlas_uid} name: {corrected_atlas_name}")

    return corrected_atlas_uid, corrected_atlas_name